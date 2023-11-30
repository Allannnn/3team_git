from slack_sdk import WebClient
from slack_sdk.errors import SlackApiError
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from dotenv import load_dotenv
import os
import re
from collections import Counter
import pandas as pd
import openpyxl

# Slack channel to send the message to
SLACK_API_TOKEN = "xoxb-6253138637332-6243077942519-9yBMEnwdIgBRMOmQ3anhRQOA"

def sendSlackWebhook(file_path):
    client = WebClient(token=SLACK_API_TOKEN)
    try:
        response = client.files_upload(
            channels="#python-test",
            file=file_path,
            title=f"위험정보 포함 파일입니다."
        )
        print(f"정상적으로 보냄")
    except SlackApiError as e:
        print(f"오류 발생 {e}")
        
load_dotenv()
SECRET_ID = os.getenv('ID')
SECRET_PASS = os.getenv('PASS')

smtp= smtplib.SMTP('smtp.naver.com',587)
smtp.ehlo()
smtp.starttls()

smtp.login(SECRET_ID,SECRET_PASS)

myemail = 'ggggame93@naver.com'
youremail = 'ggggame93@naver.com'

msg = MIMEMultipart()

msg['Subject'] ="위험정보 포함 파일 입니다."
msg['From'] = myemail
msg['To'] = youremail

text ="""
        <html>
        <body>
        <h2>위험 파일 입니다.</h2>
        </body>
        </html>
    """
    
contentPart = MIMEText(text,'html')
msg.attach(contentPart)

file_name='2023-11-30_insert_member.xlsx'
etc_file_path = fr'{file_name}'
with open (etc_file_path,'rb') as f:
    etc_part = MIMEApplication(f.read())
    etc_part.add_header('Content-Disposition','attachment', filename=etc_file_path)
    msg.attach(etc_part)
        
smtp.sendmail(myemail,youremail,msg.as_string())
smtp.quit()

#for file_name in files:
output_path = f"{file_name}"
sendSlackWebhook(output_path)
if file_name.endswith('.log'):
    with open(file_name,'r') as f:
        ip_list = re.findall(r"\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}", str(f.read()))
    ip_counter = Counter(ip_list)
    top_10_ips = ip_counter.most_common(10)
    print(top_10_ips)

wb = openpyxl.load_workbook(file_name)
sheet = wb.active

info_warning = False
phone_pattern = r'\d{3}-\d{3,4}-\d{4}'
email_pattern = r"[a-zA-Z0-9._+-]+@[a-zA-Z0-9]+\.[a-zA-Z]{2,4}"
for row in sheet.iter_rows():
    for cell in row:
        if re.findall(phone_pattern,str(cell.value)):
            info_warning = True
        elif re.findall(email_pattern,str(cell.value)):
            info_warning = True
if info_warning:
    print("위험 정보 있음")
else:
    print("위험 정보 없음")

for row in sheet.iter_rows():
    for cell in row:
        print(str(cell))